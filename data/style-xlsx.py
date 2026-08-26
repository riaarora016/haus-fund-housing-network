#!/usr/bin/env python3
# Styles data/exports/biopunk-housing-tracker.xlsx in place: dark header band, frozen panes,
# tier colors on the Priority list, autofilters, readable text tabs. Run after export-xlsx.ts.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PATH = 'data/exports/biopunk-housing-tracker.xlsx'
wb = openpyxl.load_workbook(PATH)
GREY = Font(color='666666'); BOLD = Font(bold=True)
HEAD = PatternFill('solid', fgColor='1F2937'); HEADF = Font(bold=True, color='FFFFFF', size=10)
TIER = {'Tier 1': 'D1FAE5', 'Tier 2': 'FEF3C7', 'Tier 3': 'FFF7ED', 'Tier 4': 'F3F4F6', 'BASELINE': 'DBEAFE', 'Parked': 'F3F4F6', 'Dead': 'FEE2E2'}
thin = Border(bottom=Side(style='hair', color='DDDDDD'))
TEXT_TABS = ('Start here', 'Scoring explained')
HOUSE_TABS = ('Punkhaus', 'Femhaus', 'Alumhaus')

def style_header(ws, row):
    for c in ws[row]:
        if c.value not in (None, ''):
            c.fill = HEAD; c.font = HEADF; c.alignment = Alignment(vertical='center')
    ws.freeze_panes = f'A{row+1}'

for name in wb.sheetnames:
    ws = wb[name]
    if name in TEXT_TABS:
        ws['A1'].font = Font(bold=True, size=16)
        for r in range(2, ws.max_row + 1):
            for c in ws[r]:
                c.alignment = Alignment(wrap_text=True, vertical='top')
            head = str(ws.cell(row=r, column=1).value or '')
            if head.isupper() or head.startswith(('THE JOB', 'DATES', 'SCORING', 'ON EVERY CALL', 'HONESTY', 'WORKED EXAMPLE', 'TABS', '1.', '2.', '3.', '4.')):
                ws.cell(row=r, column=1).font = BOLD
        continue
    a1 = str(ws.cell(row=1, column=1).value or '')
    b1 = ws.cell(row=1, column=2).value
    # a note line is a single long cell in A1 with nothing beside it
    header_row = 2 if (a1.startswith('NOTE:') or (a1 and b1 in (None, ''))) else 1
    if header_row == 2:
        ws.cell(row=1, column=1).font = Font(italic=True, color='92400E')
        ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[1].height = 56
    style_header(ws, header_row)
    for r in range(header_row + 1, ws.max_row + 1):
        for c in ws[r]:
            c.border = thin
            c.alignment = Alignment(vertical='top', wrap_text=(name == 'Templates'))
    if name == 'Priority list':
        for r in range(header_row + 1, ws.max_row + 1):
            tier = str(ws.cell(row=r, column=2).value or '')
            for key, color in TIER.items():
                if tier.startswith(key):
                    fill = PatternFill('solid', fgColor=color)
                    for col in range(1, 4):
                        ws.cell(row=r, column=col).fill = fill
                    break
            ws.cell(row=r, column=3).font = BOLD
            if tier.startswith(('Dead', 'Parked')):
                for c in ws[r]:
                    c.font = GREY
        ws.auto_filter.ref = f'A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}'
    if name in HOUSE_TABS:
        for r in range(header_row + 1, ws.max_row + 1):
            first = str(ws.cell(row=r, column=1).value or '')
            ws.cell(row=r, column=2).font = BOLD
            if first == '' and str(ws.cell(row=r, column=2).value or '').startswith('---'):
                for c in ws[r]:
                    c.font = Font(italic=True, color='92400E')
                ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor='FEF3C7')
        # a real table (banded rows, filter buttons); imports into Google Sheets as a proper table
        ref = f'A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}'
        t = Table(displayName=name.replace(' ', ''), ref=ref)
        t.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(t)
    if name == 'Templates':
        for r in range(header_row + 1, ws.max_row + 1):
            ws.row_dimensions[r].height = 150
wb.save(PATH)
print('styled', wb.sheetnames)
